"""
UTILIDADES PARA EXPORTACION DE REPORTES
"""
import os
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
from datetime import datetime

# ══════════════════════════════════════════════════════════
#  RUTAS DE LOGO
# ══════════════════════════════════════════════════════════

LOGO_PATH = 'file:///' + os.path.join(
    settings.BASE_DIR, 'app', 'static', 'imagenes', 'logo empresa.jpeg'
).replace('\\', '/')

LOGO_PATH_EXCEL = os.path.join(
    settings.BASE_DIR, 'app', 'static', 'imagenes', 'logo empresa.jpeg'
)


# ══════════════════════════════════════════════════════════
#  COLORES CORPORATIVOS — Acerautos
#  El ROJO es dominante, igual que en el PDF
# ══════════════════════════════════════════════════════════

COLOR_ROJO        = 'C0392B'   # Rojo Acerautos   — header + encabezados tabla
COLOR_ROJO_TITULO = 'A93226'   # Rojo más oscuro  — banner título
COLOR_AZUL        = '1A2F5A'   # Azul oscuro      — acento en pie
COLOR_ACENTO      = 'FDECEA'   # Rosa muy pálido  — filas alternas
COLOR_BLANCO      = 'FFFFFF'
COLOR_TEXTO       = '2C2C2C'   # Casi negro       — texto de datos
COLOR_TEXTO_META  = '922B21'   # Rojo oscuro      — textos secundarios


# ══════════════════════════════════════════════════════════
#  EXPORTAR PDF
# ══════════════════════════════════════════════════════════

def exportar_pdf(titulo, columnas, datos, nombre_archivo):
    contexto = {
        'titulo': titulo,
        'columnas': columnas,
        'datos': datos,
        'logo_url': LOGO_PATH,
    }
    html_string = render_to_string('reportes/reporte_pdf.html', contexto)
    html_object = HTML(string=html_string, base_url='.')
    pdf_bytes = html_object.write_pdf()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    return response


# ══════════════════════════════════════════════════════════
#  EXPORTAR EXCEL  — con logo, colores Acerautos, generado por
# ══════════════════════════════════════════════════════════

def exportar_excel(titulo, columnas, datos, nombre_archivo, generado_por=None):
    workbook  = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"

    num_cols   = len(columnas)
    ultima_col = get_column_letter(num_cols)

    fecha_texto     = datetime.now().strftime('%d/%m/%Y  %H:%M')
    generador_texto = str(generado_por) if generado_por else 'Sistema'

    # ── Fills ──────────────────────────────────────────────
    fill_rojo        = PatternFill(start_color=COLOR_ROJO,        end_color=COLOR_ROJO,        fill_type='solid')
    fill_rojo_titulo = PatternFill(start_color=COLOR_ROJO_TITULO, end_color=COLOR_ROJO_TITULO, fill_type='solid')
    fill_acento      = PatternFill(start_color=COLOR_ACENTO,      end_color=COLOR_ACENTO,      fill_type='solid')
    fill_blanco      = PatternFill(start_color=COLOR_BLANCO,      end_color=COLOR_BLANCO,      fill_type='solid')
    fill_pie         = PatternFill(start_color='F4F4F4',          end_color='F4F4F4',          fill_type='solid')

    # ── Bordes ─────────────────────────────────────────────
    sin_borde  = Border()
    borde_fino = Border(
        left=Side(style='thin',   color='E0E0E0'),
        right=Side(style='thin',  color='E0E0E0'),
        top=Side(style='thin',    color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0'),
    )
    borde_header = Border(
        left=Side(style='thin',   color='E8908A'),
        right=Side(style='thin',  color='E8908A'),
        top=Side(style='thin',    color='E8908A'),
        bottom=Side(style='thin', color='E8908A'),
    )
    borde_pie = Border(top=Side(style='medium', color=COLOR_ROJO))

    # ══════════════════════════════════════════════════════
    #  FILA 1 — Header rojo: [Logo] · ACERAUTOS · Fecha/Hora
    # ══════════════════════════════════════════════════════
    worksheet.row_dimensions[1].height = 58

    for col in range(1, num_cols + 1):
        c = worksheet.cell(row=1, column=col)
        c.fill   = fill_rojo
        c.border = sin_borde

    if os.path.exists(LOGO_PATH_EXCEL):
        try:
            logo_img        = XLImage(LOGO_PATH_EXCEL)
            logo_img.height = 52
            logo_img.width  = 115
            logo_img.anchor = 'A1'
            worksheet.add_image(logo_img)
        except Exception:
            pass

    col_emp_fin = max(num_cols - 1, 2)
    if col_emp_fin > 2:
        worksheet.merge_cells(f'B1:{get_column_letter(col_emp_fin)}1')
    cel_emp           = worksheet.cell(row=1, column=2)
    cel_emp.value     = 'ACERAUTOS'
    cel_emp.font      = Font(name='Calibri', size=18, bold=True, color=COLOR_BLANCO)
    cel_emp.fill      = fill_rojo
    cel_emp.alignment = Alignment(horizontal='center', vertical='center')
    cel_emp.border    = sin_borde

    fecha_parts     = fecha_texto.split('  ')
    cel_fecha       = worksheet.cell(row=1, column=num_cols)
    cel_fecha.value = f'Fecha: {fecha_parts[0]}\nHora:  {fecha_parts[1] if len(fecha_parts) > 1 else ""}'
    cel_fecha.font  = Font(name='Calibri', size=9, color=COLOR_BLANCO)
    cel_fecha.fill  = fill_rojo
    cel_fecha.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    cel_fecha.border    = sin_borde

    # ══════════════════════════════════════════════════════
    #  FILA 2 — Banner título (rojo más oscuro)
    # ══════════════════════════════════════════════════════
    worksheet.row_dimensions[2].height = 28
    worksheet.merge_cells(f'A2:{ultima_col}2')
    cel_titulo           = worksheet['A2']
    cel_titulo.value     = titulo.upper()
    cel_titulo.font      = Font(name='Calibri', size=14, bold=True, color=COLOR_BLANCO)
    cel_titulo.fill      = fill_rojo_titulo
    cel_titulo.alignment = Alignment(horizontal='center', vertical='center')
    cel_titulo.border    = sin_borde

    # ══════════════════════════════════════════════════════
    #  FILA 3 — Metadatos: Generado por
    # ══════════════════════════════════════════════════════
    worksheet.row_dimensions[3].height = 17

    mitad = max(num_cols // 2, 1)
    worksheet.merge_cells(f'A3:{get_column_letter(mitad)}3')
    cel_izq        = worksheet['A3']
    cel_izq.fill   = fill_blanco
    cel_izq.border = sin_borde

    col_gen = mitad + 1
    if col_gen <= num_cols:
        worksheet.merge_cells(f'{get_column_letter(col_gen)}3:{ultima_col}3')
    cel_gen           = worksheet.cell(row=3, column=min(col_gen, num_cols))
    cel_gen.value     = f'Generado por: {generador_texto}'
    cel_gen.font      = Font(name='Calibri', size=9, italic=True, color=COLOR_TEXTO_META)
    cel_gen.fill      = fill_blanco
    cel_gen.alignment = Alignment(horizontal='right', vertical='center')
    cel_gen.border    = sin_borde

    # ══════════════════════════════════════════════════════
    #  FILA 4 — Encabezados (ROJO + texto blanco)
    # ══════════════════════════════════════════════════════
    worksheet.row_dimensions[4].height = 22

    for col_num, columna in enumerate(columnas, 1):
        cell           = worksheet.cell(row=4, column=col_num)
        cell.value     = str(columna).upper()
        cell.font      = Font(name='Calibri', size=10, bold=True, color=COLOR_BLANCO)
        cell.fill      = fill_rojo
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = borde_header

    # ══════════════════════════════════════════════════════
    #  FILAS 5+ — Datos con filas alternas blanco / rosa pálido
    # ══════════════════════════════════════════════════════
    for row_num, fila in enumerate(datos, 5):
        worksheet.row_dimensions[row_num].height = 17

        if isinstance(fila, dict):
            valores = [fila.get(col.lower().replace(' ', '_'), '') for col in columnas]
        else:
            valores = list(fila)

        fill_fila = fill_blanco if (row_num - 5) % 2 == 0 else fill_acento

        for col_num, valor in enumerate(valores, 1):
            cell           = worksheet.cell(row=row_num, column=col_num)
            cell.value     = valor
            cell.font      = Font(name='Calibri', size=10, bold=False, color='2C2C2C')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border    = borde_fino
            cell.fill      = fill_fila

    # ══════════════════════════════════════════════════════
    #  PIE DE PÁGINA — separador rojo + fondo gris
    # ══════════════════════════════════════════════════════
    fila_pie        = worksheet.max_row + 2
    total_registros = len(datos)
    worksheet.row_dimensions[fila_pie].height = 18

    t1 = max(num_cols // 3, 1)
    t2 = t1 + 1
    t3 = max((num_cols * 2) // 3, t2)
    t4 = t3 + 1

    worksheet.merge_cells(f'A{fila_pie}:{get_column_letter(t1)}{fila_pie}')
    c1           = worksheet[f'A{fila_pie}']
    c1.value     = f'Total de registros: {total_registros}'
    c1.font      = Font(name='Calibri', size=9, color=COLOR_TEXTO_META)
    c1.fill      = fill_pie
    c1.alignment = Alignment(horizontal='left', vertical='center')
    c1.border    = borde_pie

    if t3 >= t2:
        worksheet.merge_cells(f'{get_column_letter(t2)}{fila_pie}:{get_column_letter(t3)}{fila_pie}')
    c2           = worksheet.cell(row=fila_pie, column=t2)
    c2.value     = 'Acerautos — Centro Integral Automotriz'
    c2.font      = Font(name='Calibri', size=9, bold=True, color=COLOR_TEXTO_META)
    c2.fill      = fill_pie
    c2.alignment = Alignment(horizontal='center', vertical='center')
    c2.border    = borde_pie

    if t4 <= num_cols:
        worksheet.merge_cells(f'{get_column_letter(t4)}{fila_pie}:{ultima_col}{fila_pie}')
    c3           = worksheet.cell(row=fila_pie, column=min(t4, num_cols))
    c3.value     = f'Generado el: {fecha_texto}'
    c3.font      = Font(name='Calibri', size=9, color=COLOR_TEXTO_META)
    c3.fill      = fill_pie
    c3.alignment = Alignment(horizontal='right', vertical='center')
    c3.border    = borde_pie

    # ══════════════════════════════════════════════════════
    #  Ajuste automático de ancho de columnas
    # ══════════════════════════════════════════════════════
    for col_num, columna in enumerate(columnas, 1):
        col_letra  = get_column_letter(col_num)
        max_length = len(str(columna))
        for row in worksheet.iter_rows(
            min_row=4, max_row=worksheet.max_row,
            min_col=col_num, max_col=col_num
        ):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
        worksheet.column_dimensions[col_letra].width = min(max_length + 4, 50)

    worksheet.column_dimensions['A'].width = max(
        worksheet.column_dimensions['A'].width, 18
    )

    # ══════════════════════════════════════════════════════
    #  Respuesta HTTP
    # ══════════════════════════════════════════════════════
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    workbook.save(response)
    return response